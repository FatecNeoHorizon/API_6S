import pandas as pd
from pathlib import Path
from collections.abc import Iterator
from itertools import chain
from openpyxl import load_workbook

CHUNK_SIZE = 10_000
HEADER_SCAN_LIMIT = 25


def _non_empty_count(row: tuple) -> int:
    return sum(1 for cell in row if cell is not None and str(cell).strip() != "")


def _build_headers(raw_header: tuple) -> list[str]:
    last_non_empty_idx = -1
    for idx, value in enumerate(raw_header):
        if value is not None and str(value).strip() != "":
            last_non_empty_idx = idx

    if last_non_empty_idx < 0:
        raise ValueError("Cabeçalho inválido no arquivo de perdas de energia.")

    header_slice = raw_header[: last_non_empty_idx + 1]
    headers: list[str] = []
    seen: dict[str, int] = {}

    for idx, value in enumerate(header_slice, start=1):
        base_name = str(value).strip() if value is not None else ""
        if not base_name:
            base_name = f"column_{idx}"

        if base_name in seen:
            seen[base_name] += 1
            col_name = f"{base_name}_{seen[base_name]}"
        else:
            seen[base_name] = 1
            col_name = base_name

        headers.append(col_name)

    return headers


def extract_losses(path: Path, chunk_size: int = CHUNK_SIZE) -> Iterator[tuple[pd.DataFrame, str]]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active

    row_iter = ws.iter_rows(values_only=True)
    scanned_rows: list[tuple] = []

    for _ in range(HEADER_SCAN_LIMIT):
        try:
            scanned_rows.append(next(row_iter))
        except StopIteration:
            break

    if not scanned_rows:
        wb.close()
        return

    header_idx = max(range(len(scanned_rows)), key=lambda idx: _non_empty_count(scanned_rows[idx]))
    header_row = scanned_rows[header_idx]

    if _non_empty_count(header_row) <= 1:
        wb.close()
        raise ValueError("Não foi possível detectar cabeçalho válido no arquivo de perdas de energia.")

    headers = _build_headers(header_row)
    width = len(headers)

    def normalize_row(row: tuple) -> tuple:
        sliced = row[:width]
        if len(sliced) < width:
            return sliced + (None,) * (width - len(sliced))
        return sliced

    data_iter = chain(scanned_rows[header_idx + 1 :], row_iter)
    buffer: list[tuple] = []

    try:
        for row in data_iter:
            normalized = normalize_row(row)
            if _non_empty_count(normalized) == 0:
                continue

            buffer.append(normalized)
            if len(buffer) >= chunk_size:
                yield pd.DataFrame(buffer, columns=headers), str(path)
                buffer.clear()

        if buffer:
            yield pd.DataFrame(buffer, columns=headers), str(path)
    finally:
        wb.close()


def extract_losses_preview(path: Path, limit: int = 50) -> list[dict]:
    preview: list[dict] = []
    for chunk_df, _source_file in extract_losses(path):
        remaining = limit - len(preview)
        if remaining <= 0:
            break
        preview.extend(chunk_df.head(remaining).to_dict(orient="records"))
    return preview[:limit]